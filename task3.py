# %%
# 为兼容低版本python（不支持函数的参数类别声明），导入__future__模块
from __future__ import annotations

# %% md
# 语法分析
# %%
import builtins
import importlib
import os

import my_sdt  # 导入语法文件
import my_sdt_action  # 导入语义动作文件

# 刷新模块, 防止修改文件后变量没有更新
importlib.reload(my_sdt)
importlib.reload(my_sdt_action)


# %%
# 重定向输出构造器
def print_redirect_builder(file_path, use_cache=True) -> callable:
    """
    重定向输出构造器，返回重定向过的输出函数
    :param file_path: 文件路径(如果空字符串那么输出到控制台，否则输出到文件)
    :param use_cache: 是否使用缓存(默认使用缓存)(如果使用缓存，那么只有flush=True时才会输出，否则输出到内部的字符串)
    :return: 重定向输出函数
    """
    output_cache = ''  # 输出缓存在字符串中

    def print_redirect(*args, mode="a", flush=not use_cache, **kwargs):
        """
        重定向输出函数
        :param args: 其他args
        :param mode: 输出模式
        :param flush: 是否刷新缓存，默认值取决于构造时的use_cache的值
        :param kwargs: 其他kwargs
        """
        nonlocal output_cache
        if not flush:  # 如果不刷新，那么只输出到缓存字符串
            sep = kwargs.get('sep', ' ')
            end = kwargs.get('end', '\n')
            output_cache += sep.join(map(str, args)) + end
        else:  # 如果刷新，那么输出到文件或控制台
            if file_path != '':
                # 打开文件并写入，如果文件不存在则创建
                with open(file_path, mode, encoding='GBK') as f:
                    if output_cache != '':
                        builtins.print(output_cache, file=f, end='', **kwargs)
                    builtins.print(*args, file=f, **kwargs)
            else:  # 输出到控制台
                if output_cache != '':
                    builtins.print(output_cache, end='', **kwargs)
                builtins.print(*args, **kwargs)

    return print_redirect


# %%
# 检查文法变量的格式
def check_grammar(_augmented_grammar: dict[str:list[my_sdt.SDT_right[str]]]):
    """
    检查文法格式是否正确，不正确则抛出断言异常
    :param _augmented_grammar: 增广文法
    """
    assert isinstance(_augmented_grammar, dict)
    for key in _augmented_grammar.keys():
        assert isinstance(key, str)
        assert isinstance(_augmented_grammar[key], list)
        for right in _augmented_grammar[key]:
            assert isinstance(right, list)
            for symbol in right:
                assert isinstance(symbol, str)
    assert "S'" in _augmented_grammar.keys()
    assert len(_augmented_grammar["S'"]) == 1
    assert len(_augmented_grammar["S'"][0]) == 1
    # 检查SDT（每个产生式中不能有重复的非终结符）
    for left_symbol in _augmented_grammar.keys():
        for right_symbols in _augmented_grammar[left_symbol]:
            # 这行产生式的所有符号
            production_symbol_list = [left_symbol] + right_symbols.rights
            # 这行的所有非终结符
            production_nonterminal_symbol_list = list(filter(
                lambda x: x in _augmented_grammar.keys(),
                production_symbol_list))
            # 如果这行的所有非终结符有重复，那么抛出异常
            assert len(production_nonterminal_symbol_list) == len(set(production_nonterminal_symbol_list)), \
                f"文法中产生式{left_symbol}->{right_symbols}中有重复的非终结符"


# %%
# 从文法中获取符号列表、非终结符列表、终结符列表
def get_symbol_list(_augmented_grammar):
    """
    从文法中获取符号列表、非终结符列表、终结符列表
    :param _augmented_grammar: 增广文法
    :return: 符号列表、非终结符列表、终结符列表
    """
    _nonterminal_symbol_list = set(_augmented_grammar.keys())
    _terminal_symbol_list = \
        (
                set([symbol for right_list in _augmented_grammar.values() for right in right_list for symbol in right])
                - _nonterminal_symbol_list
        ) | {'$'}
    _nonterminal_symbol_list = sorted(list(_nonterminal_symbol_list),
                                      key=lambda x: (x != "S'", x != _augmented_grammar["S'"][0][0], x))
    _terminal_symbol_list = sorted(list(_terminal_symbol_list), key=lambda x: (x == '$', x))
    _symbol_list = _nonterminal_symbol_list + _terminal_symbol_list
    return _symbol_list, _nonterminal_symbol_list, _terminal_symbol_list


# %%
# 将文法按顺序编号，便于归约时输出r1r2r3....
def get_augmented_grammar_to_index(_augmented_grammar: dict[str:list[my_sdt.SDT_right[str]]]) -> \
        tuple[dict[tuple[str, tuple[str]]:int], list[tuple[str, my_sdt.SDT_right[str]]]]:
    """
    将文法按顺序编号，便于归约时输出r1r2r3....
    :param _augmented_grammar: 增广文法
    :return: 文法序号字典、排序后的文法列表
    """
    _augmented_grammar_to_index = dict()
    _augmented_grammar_order_list = []
    index = 0
    for key in _augmented_grammar.keys():
        for right in _augmented_grammar[key]:
            _augmented_grammar_to_index[(key, tuple(right))] = index
            index += 1
            _augmented_grammar_order_list.append((key, right))

    return _augmented_grammar_to_index, _augmented_grammar_order_list


# %%
# 项集类，每一个实例都是一个项集
class ItemSet:
    def __init__(self, item_set: set | list):
        """
        初始化项集
        :param item_set: 项集
        """
        # 项集，使用set变量便于进行集合运算
        self.item_set = set(item_set)
        # 项集(有序)，使用list变量可以按顺序遍历(set是无序的，所以加了一个list变量，内容一样)
        self.item_set_order_list = list(self.item_set)
        for item in self.item_set:
            assert isinstance(item, Item)

    def add(self, item):
        """
        添加项，同时更新set和list
        :param item: 项
        """
        assert isinstance(item, Item)
        self.item_set.add(item)
        self.item_set_order_list.append(item)

    def __len__(self):
        """返回长度"""
        return len(self.item_set)

    def __eq__(self, other):
        """判断是否相等"""
        assert isinstance(other, ItemSet)
        return self.item_set == other.item_set

    def hash_str(self):
        """便于hash，hash时根据字符串产生hashID来判断是否相等"""
        return str(
            ','.join(
                list(map(
                    lambda item: item.hash_str(),
                    list(sorted(
                        self.item_set, key=lambda item:
                        (item.left_symbol, item.right_symbol_list, item.dot_index)
                    ))
                ))
            )
        )

    def __hash__(self):
        """便于set去重，set会根据字符串产生hashID来判断是否相等，相等时调用__eq__判断是否真的相等"""
        return hash(self.hash_str())

    def __str__(self):
        """输出字符串"""
        return ' \n'.join(list(map(str, self.item_set_order_list)))


# 项类，每一个实例都是一个项
class Item:
    def __init__(self, left_symbol: str, right_symbol_list: list, dot_index: int):
        """
        初始化项集, 例如 E->E·+T, 写作Item('E', ['E','+','T'], 1)
        :param left_symbol: 产生式左边的符号，如'E'
        :param right_symbol_list: 产生式右边的符号(列表)，如['E','+','T']
        :param dot_index: 点的位置，dot_index ∈ [0,len(right_symbol_list)]，如1
        """
        self.left_symbol = left_symbol
        self.right_symbol_list = right_symbol_list
        self.dot_index = dot_index

    def __eq__(self, other):
        """判断是否相等"""
        return self.left_symbol == other.left_symbol \
            and self.right_symbol_list == other.right_symbol_list \
            and self.dot_index == other.dot_index

    def hash_str(self):
        """便于hash，hash时根据字符串产生hashID来判断是否相等"""
        return str((self.left_symbol, self.right_symbol_list, self.dot_index))

    def __hash__(self):
        """便于set去重，set会根据字符串产生hashID来判断是否相等，相等时调用__eq__判断是否真的相等"""
        return hash(self.hash_str())

    def __str__(self):
        """输出字符串"""
        return self.left_symbol + ' -> ' + \
            ' '.join(self.right_symbol_list[:self.dot_index] + ['·'] + self.right_symbol_list[self.dot_index:])


# %%
# 根据增广文法中的信息，给定一个项集，返回这个项集的闭包
def get_closure(_augmented_grammar: dict[str:list[list[str]]], itemSet: ItemSet) -> ItemSet:
    """
    根据增广文法中的信息，给定一个项集，返回这个项集的闭包，和PPT中的算法一样，只不过将每次待计算的项加入队列，而非每次遍历直到项集不再增加
    :param _augmented_grammar: 增广文法
    :param itemSet: 项集
    :return: 这个项集的闭包
    """
    itemSet_result = ItemSet(itemSet.item_set_order_list.copy())
    item_queue = itemSet.item_set_order_list.copy()  # 待处理的项(队列)
    while len(item_queue) > 0:  # 队列不为空
        item = item_queue.pop(0)  # 取出第一个项
        # 如果项集中的点不在最后，展开
        if item.dot_index < len(item.right_symbol_list):
            # 待展开的符号
            symbol = item.right_symbol_list[item.dot_index]
            # 如果是非终结符
            if symbol in _augmented_grammar.keys():
                # 展开
                for right in _augmented_grammar[symbol]:
                    new_item = Item(symbol, right, 0)
                    if new_item not in itemSet_result.item_set:
                        itemSet_result.add(new_item)
                        item_queue.append(new_item)
    return itemSet_result


# %%
# 根据增广文法中的信息，给定一个项集和一个符号，返回这个项集关于这个符号的goto
def get_goto(_augmented_grammar: dict[str:list[list[str]]], itemSet: ItemSet, X: str) -> ItemSet:
    """
    获取项集I关于符号X的goto，和PPT中的算法一致
    :param _augmented_grammar: 增广文法
    :param itemSet: 项集
    :param X: 符号
    :return: 项集I关于符号X的goto
    """
    J = ItemSet(list())
    for item in itemSet.item_set_order_list:
        if item.dot_index < len(item.right_symbol_list) and item.right_symbol_list[item.dot_index] == X:
            new_item = Item(item.left_symbol, item.right_symbol_list, item.dot_index + 1)
            J.add(new_item)
    return get_closure(_augmented_grammar, J)


# %%
# 根据增广文法中的信息，获取项集族(所有项集)
def get_itemSetFamily(_augmented_grammar: dict[str:list[list[str]]], _symbol_list: list[str]) -> list[ItemSet]:
    """
    获取项集族(所有项集)，和PPT中的算法一致，只不过将每次待计算的项集合加入队列，而非每次遍历直到项集族不再增加
    :param _augmented_grammar: 增广文法
    :param _symbol_list: 符号列表
    :return: 项集族(所有项集)
    """
    # 用于存储所有项集，使用set避免重复(set内部使用hash方法，我们已经在ItemSet类中重写了hash方法)
    itemSet_set = set()
    itemSet_set_order_list = list()
    # 初始化
    I0 = get_closure(_augmented_grammar, ItemSet({Item("S'", _augmented_grammar["S'"][0], 0)}))
    itemSet_set.add(I0)
    itemSet_set_order_list.append(I0)
    # 待处理的项集
    itemSet_queue = [I0]
    while len(itemSet_queue) > 0:
        itemSet = itemSet_queue.pop(0)
        for X in _symbol_list:
            J = get_goto(_augmented_grammar, itemSet, X)
            if len(J) > 0 and J not in itemSet_set:
                itemSet_set.add(J)
                itemSet_set_order_list.append(J)
                itemSet_queue.append(J)
    return itemSet_set_order_list


# %%
# 从增广文法中获取First集
def get_first(_augmented_grammar: dict[str:list[list[str]]], _nonterminal_symbol_list: list[str],
              _terminal_symbol_list: list[str]) -> dict[str:set[str]]:
    """
    获取First集，和PPT中的算法一致
    :param _augmented_grammar: 增广文法
    :param _nonterminal_symbol_list: 非终结符列表
    :param _terminal_symbol_list: 终结符列表
    :return: First集
    """
    _first = dict()
    for X in _nonterminal_symbol_list:
        _first[X] = set()
    for X in _terminal_symbol_list:
        _first[X] = {X}
    # 用于判断First集是否改变
    is_changed = True
    while is_changed:
        is_changed = False
        for X in _nonterminal_symbol_list:
            for right in _augmented_grammar[X]:
                if len(right) == 0:  # X->ε
                    if 'ε' not in _first[X]:
                        _first[X].add('ε')
                        is_changed = True
                else:
                    for symbol in right:
                        if symbol in _nonterminal_symbol_list:  # X->Y...
                            if _first[symbol] - _first[X] != set():
                                _first[X] = _first[X] | _first[symbol]
                                is_changed = True
                        elif symbol in _terminal_symbol_list:
                            if symbol not in _first[X]:
                                _first[X].add(symbol)
                                is_changed = True
                        if 'ε' not in _first[symbol]:
                            break
    return _first


def get_compound_first(compound_symbol_list: list[str], _first: dict[str, set | set[str]]) -> set[str]:
    """
    获取多个符号的First集，如果其中有一个符号的First集包含ε，那么就要继续往后找，直到找到一个不包含ε的符号的First集
    :return: 多个符号的First集
    """
    compound_first = set()
    for i in range(len(compound_symbol_list)):
        compound_first = compound_first | _first[compound_symbol_list[i]] - {'ε'}
        if 'ε' not in _first[compound_symbol_list[i]]:
            break
        if i == len(compound_symbol_list) - 1:
            compound_first.add('ε')
    return compound_first


# %%
# 从增广文法中获取Follow集
def get_follow(_augmented_grammar: dict[str:list[list[str]]], _nonterminal_symbol_list: list[str],
               _terminal_symbol_list: list[str],
               _first: dict[str, set | set[str]]):
    """
    获取Follow集，和PPT中的算法一致
    1)将$放到FOLLOW(S')中,其中S'是开始符号,而$是输入右端的结束标记。
    2)如果存在一个产生式A→aBβ，那么FIRST(β)中除ε之外的所有符号都在FOLLOW(B)中。
    3）如果存在一个产生式A→aB，或存在产生式A→aBβ且FIRST(β)包含ε,那么FOLLOW(A)中的所有符号都在FOLLOW(B)中,
    :param _augmented_grammar: 增广文法
    :param _nonterminal_symbol_list: 非终结符列表
    :param _terminal_symbol_list: 终结符列表
    :param _first: First集
    :return: Follow集
    """
    _follow = dict()
    for A in _nonterminal_symbol_list:
        _follow[A] = set()
    # 1)将$放到FOLLOW(S')中,其中S'是开始符号,而$是输入右端的结束标记。
    _follow["S'"] = {'$'}
    # 2)如果存在一个产生式A→αBβ，那么FIRST(β)中除ε之外的所有符号都在FOLLOW(B)中。
    # 遍历所有产生式，将形如A->αBβ的β的FIRST集合加入FOLLOW(B)
    for A in _nonterminal_symbol_list:
        for right in _augmented_grammar[A]:
            if len(right) == 0:
                continue
            for i in range(len(right) - 1):
                B = right[i]
                beta = right[i + 1:]
                if B in _nonterminal_symbol_list:
                    _follow[B] = _follow[B] | get_compound_first(beta, _first) - {'ε'}
    # 3）如果存在一个产生式A→αB，或存在产生式A→αBβ且FIRST(β)包含ε,那么FOLLOW(A)中的所有符号都在FOLLOW(B)中,
    # 4）如果A->αBC1C2...Cn，且C1->·，那么Follow(C1)全部都在Follow(B)中，如果又有C2->·，那么Follow(C2)全部都在Follow(B)中，直到n
    # 遍历所有产生式，将形如A->αB的FOLLOW(A)加入FOLLOW(B)，或者将形如A->αBβ且FIRST(β)包含ε的FOLLOW(A)加入FOLLOW(B)
    # 未变
    is_changed = True
    while is_changed:
        is_changed = False
        for A in _nonterminal_symbol_list:
            for right in _augmented_grammar[A]:
                if len(right) == 0:
                    continue
                for i in range(len(right)):
                    B = right[i]
                    if B in _nonterminal_symbol_list:  # 非终结符才需要考虑，设为B
                        if i == len(right) - 1:  # B位于产生式最后，即A->αB（3.1的情况）
                            if _follow[A] - _follow[B] != set():
                                _follow[B] = _follow[B] | _follow[A]
                                is_changed = True
                        else:
                            beta = right[i + 1:]  # B不位于产生式最后，即A->αBβ（3.2的情况）
                            if 'ε' in get_compound_first(beta, _first):  # FIRST(β)包含ε
                                if _follow[A] - _follow[B] != set():
                                    _follow[B] = _follow[B] | _follow[A]
                                    is_changed = True
                            # 判断（4）的情况
                            for ii in range(i + 1, len(right)):
                                C = right[ii]
                                if C not in _nonterminal_symbol_list:
                                    break
                                if list() not in _augmented_grammar[C]:
                                    break
                                if _follow[C] - _follow[B] != set():
                                    _follow[B] = _follow[B] | _follow[C]
                                    is_changed = True
    return _follow


# %%
# 给定增广文法，根据上面求出的项集族、First集、Follow集，获取SLR分析表
def get_SLR_table(_augmented_grammar: dict[str:list[list[str]]],
                  _augmented_grammar_to_index: dict[tuple[str, tuple[str]]:int],
                  _symbol_list: list[str], _nonterminal_symbol_list: list[str], _terminal_symbol_list: list[str],
                  _first: dict[str, set | set[str]], _follow: dict[str, set | set[str]],
                  _itemSetFamily: list[ItemSet]) -> \
        tuple[dict[int:dict[str:str]], dict[int:dict[str:str]]]:
    """
    获取SLR分析表\n
    格式样例：\n
    itemSetFamily = [I0, I1, I2] 每个Ii都是一个ItemSet\n
    action = {0: {'i': 's1', '+': 's2'  '$': ''   },\n
              1: {'i': ''  , '+': 'r1', '$': 'acc'},\n
              2: {'i': 'r2', '+': 's2', '$': 'r2' }}\n
    goto = {0: {'E': '1', 'T': '2'},\n
            1: {'E': '',  'T': '' },\n
            2: {'E': '',  'T': '' }}\n
    action每个格子都有值，缺省为''   \n
    goto每个格子都有值，缺省为''   \n
    :param _augmented_grammar: 增广文法
    :param _augmented_grammar_to_index: 增广文法序号
    :param _symbol_list: 符号列表
    :param _nonterminal_symbol_list: 非终结符列表
    :param _terminal_symbol_list: 终结符列表
    :param _first: First集
    :param _follow: Follow集
    :param _itemSetFamily: 状态集(每个状态其实就是一个ItemSet)
    :return: action表、goto表
    """
    # 项集族是一个有序的list，我将其编号并存储在字典中，方便查找
    itemSetFamily_to_index = {_itemSetFamily[i]: i for i in range(len(_itemSetFamily))}
    # 将goto表初始化为空表
    _goto = dict()
    for i in range(len(_itemSetFamily)):
        _goto[i] = dict()
        for X in _nonterminal_symbol_list:
            _goto[i][X] = ""
    # 将action表初始化为空表
    _action = dict()
    for i in range(len(_itemSetFamily)):
        _action[i] = dict()
        for X in _terminal_symbol_list:
            _action[i][X] = ''
    # 对项集族中的每个项集，判断它们能通过哪个符号连接到哪个项集，记入action和goto表；再判断是否能够归约，把归约记录在action中
    for itemSet in _itemSetFamily:
        itemSet_index = itemSetFamily_to_index[itemSet]
        # 遍历文法符号，试图操作(移入和跳转)
        for X in _symbol_list:
            goto_itemSet = get_goto(_augmented_grammar, itemSet, X)
            if len(goto_itemSet) > 0:
                assert goto_itemSet in itemSetFamily_to_index.keys()
                goto_itemSet_index = itemSetFamily_to_index[goto_itemSet]
                # 如果X是非终结符，goto加一条'跳转'边
                if X in _augmented_grammar.keys():
                    _goto[itemSet_index][X] += str(goto_itemSet_index)  # FIXME 冲突时优先移入(=号而不是+=号)
                # 如果X是终结符，action加一条'移入'边
                else:
                    _action[itemSet_index][X] += 's' + str(goto_itemSet_index)  # FIXME 冲突时优先移入(=号而不是+=号)
        # 试图归约
        reduce_items = list(  # 找出所有可以归约的项
            filter(lambda _item: _item.dot_index == len(_item.right_symbol_list), itemSet.item_set_order_list))
        for reduce_item in reduce_items:  # 对于每一个可以归约的项
            if reduce_item.left_symbol == "S'":  # 如果是S'->S，那么acc
                _action[itemSet_index]['$'] += 'acc'
            else:  # 否则，找到归约的产生式，加入action表
                reduce_item_index = \
                    _augmented_grammar_to_index[(reduce_item.left_symbol, tuple(reduce_item.right_symbol_list))]
                for X in _terminal_symbol_list:
                    if X in _follow[reduce_item.left_symbol]:  # 如果X在FOLLOW(A)中，才会归约
                        _action[itemSet_index][X] += 'r' + str(reduce_item_index)  # FIXME 冲突时优先移入(=号而不是+=号)

    return _action, _goto


# %%
# 检查SLR冲突
def check_SLR(_action: dict[int:dict[str:str]], _goto: dict[int:dict[str:str]]):
    """
    检查SLR分析表是否有冲突，有冲突则抛出断言异常
    :param _action: action表
    :param _goto: goto表
    """
    for i in _action.keys():
        for X in _action[i].keys():
            if _action[i][X] != '' and _action[i][X] != 'acc':
                assert _action[i][X].count('s') + _action[i][X].count('r') < 2, \
                    f"归约归约冲突或移入归约冲突: 当I{i}遇到{X}执行{_action[i][X]}"
                assert _action[i][X] == '' or _action[i][X] == 'acc' or \
                       (_action[i][X][0] in ['s', 'r'] and _action[i][X][1:].isnumeric()), \
                    f"未知的action表动作(理论上不可能, 似乎是action表的生成存在问题): 当I{i}遇到{X}执行{_action[i][X]}"
    for i in _goto.keys():
        for X in _goto[i].keys():
            if _goto[i][X] != '':
                assert _goto[i][X].isnumeric(), f"goto表的值不是数字: 当I{i}遇到{X}跳转到{_goto[i][X]}"


# %%
# 输出SLR分析表到文本文件
def print_SLR_table(_nonterminal_symbol_list: list[str], _terminal_symbol_list: list[str],
                    _itemSetFamily: list[ItemSet], _action: dict[int:dict[str:str]], _goto: dict[int:dict[str:str]],
                    print_filepath: str = ''):
    """
    输出SLR分析表
    :param _nonterminal_symbol_list: 非终结符列表
    :param _terminal_symbol_list: 终结符列表
    :param _itemSetFamily: 状态列表
    :param _action: action表
    :param _goto: goto表
    :param print_filepath: 输出文件路径
    """
    # 重定向输出
    print = print_redirect_builder(print_filepath)
    # 输出状态
    print("itemSetFamily:\n")
    for i in range(len(_itemSetFamily)):
        print("I" + str(i) + ":")
        print(_itemSetFamily[i])
        print()

    # 准备输出SLR表
    print_table_cell = [["" for _j in range(1 + len(_terminal_symbol_list) + len(_nonterminal_symbol_list) - 1)]
                        for _i in range(1 + len(_itemSetFamily))]
    # 输出表头
    print_table_cell[0] = [''] + _terminal_symbol_list + [_ for _ in _nonterminal_symbol_list if _ != "S'"]
    # 输出表体
    for i in range(1, len(print_table_cell)):
        # 输出表的第一列
        print_table_cell[i][0] = 'I' + str(i - 1)
        # 输出action和goto
        for j in range(1, len(print_table_cell[i])):
            if print_table_cell[0][j] in _terminal_symbol_list:
                print_table_cell[i][j] = _action[i - 1][print_table_cell[0][j]]
            else:
                print_table_cell[i][j] = _goto[i - 1][print_table_cell[0][j]]

    # 将待输出内容的每列的字符串长度对齐
    max_column_length = [0 for _j in range(len(print_table_cell[0]))]
    for i in range(len(print_table_cell)):
        for j in range(len(print_table_cell[i])):
            max_column_length[j] = max(max_column_length[j], len(print_table_cell[i][j]))
    for i in range(len(print_table_cell)):
        for j in range(len(print_table_cell[i])):
            print_table_cell[i][j] = print_table_cell[i][j].ljust(max_column_length[j])

    # 输出SLR表
    for i in range(len(print_table_cell)):
        for j in range(len(print_table_cell[i])):
            print(print_table_cell[i][j], end='  ')
        print()

    print("\n", flush=True)


# %%
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# 输出SLR分析表到Excel文件
def print_SLR_table_to_excel(_nonterminal_symbol_list: list[str], _terminal_symbol_list: list[str],
                             _itemSetFamily: list[ItemSet], _action: dict[int:dict[str:str]],
                             _goto: dict[int:dict[str:str]], output_filepath: str = 'SLR_table.xlsx'):
    """
    输出SLR分析表到Excel文件
    :param _nonterminal_symbol_list: 非终结符列表
    :param _terminal_symbol_list: 终结符列表
    :param _itemSetFamily: 状态列表
    :param _action: action表
    :param _goto: goto表
    :param output_filepath: 输出Excel文件路径
    """
    # 创建一个Excel工作簿
    wb = Workbook()
    # 选择默认的工作表
    ws = wb.active

    # 设置表头第一行
    ws.append(["项集族", "状态", "ACTION"] + [""] * (len(_terminal_symbol_list) - 1) + ["GOTO"] +
              [""] * (len(_nonterminal_symbol_list) - 2))

    # 设置表头第二行
    second_header_row = [""] * 2 + _terminal_symbol_list + [_ for _ in _nonterminal_symbol_list if _ != "S'"]
    ws.append(second_header_row)

    # 输出表体
    for i in range(len(_itemSetFamily)):
        item_set_str = str(_itemSetFamily[i])
        state_str = "I" + str(i)
        # 输出action和goto
        action_row = [_action[i].get(symbol, "") for symbol in _terminal_symbol_list]
        goto_row = [_goto[i].get(symbol, "") for symbol in _nonterminal_symbol_list if symbol != "S'"]
        # 将每个元素写入单独的单元格
        body_row = [item_set_str, state_str] + action_row + goto_row
        ws.append(body_row)

    # 所有单元格设置为字符串格式，防止识别成公式或者数字之类的
    for i in range(1, len(_itemSetFamily) + 3):
        for j in range(1, len(_terminal_symbol_list) + len(_nonterminal_symbol_list) + 2):
            ws.cell(row=i, column=j).data_type = 's'

    # # 设置每一列的固定宽度
    # column_widths = [15, 15] + [10] * (len(_terminal_symbol_list) + len(_nonterminal_symbol_list) - 2)
    # for i, width in enumerate(column_widths):
    #     column_letter = get_column_letter(i + 1)
    #     ws.column_dimensions[column_letter].width = width
    # 求出每列的最大宽度
    max_column_length = [0] * (len(_terminal_symbol_list) + len(_nonterminal_symbol_list) + 2)
    for i in range(len(_itemSetFamily) + 2):
        for j in range(len(_terminal_symbol_list) + len(_nonterminal_symbol_list) + 2):
            max_column_length[j] = max(max_column_length[j],
                                       max(list(map(len, str(ws.cell(row=i + 1, column=j + 1).value).split('\n'))))
                                       )
    # 设置每一列的宽度
    for i, width in enumerate(max_column_length):
        column_letter = get_column_letter(i + 1)
        ws.column_dimensions[column_letter].width = width + 0.78  # 0.78为自动的单元格边距宽度

    # 合并部分表头单元格
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)  # 合并项集族
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)  # 合并状态
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=2 + len(_terminal_symbol_list))  # 合并ACTION
    ws.merge_cells(start_row=1, start_column=3 + len(_terminal_symbol_list), end_row=1,
                   end_column=3 + len(_terminal_symbol_list) + len(_nonterminal_symbol_list) - 2)  # 合并GOTO

    # 保存Excel文件
    wb.save(output_filepath)


# %%
# 开始推导过程
def SLR_parsing(_tokens: list[Token], _token_place: list[tuple[int, int]],
                _augmented_grammar_order_list: list[tuple[str, my_sdt.SDT_right[str]]],
                _nonterminal_symbol_list: list[str], _terminal_symbol_list: list[str],
                _itemSetFamily: list[ItemSet], _action: dict[int:dict[str:str]], _goto: dict[int:dict[str:str]]) \
        -> tuple[list[list[str]], list[list[str]]]:
    """
    SLR推导
    :param _tokens: 词法分析结果
    :param _token_place: token的位置(行号、列号)的列表
    :param _augmented_grammar_order_list: 增广文法序号列表
    :param _nonterminal_symbol_list: 非终结符列表
    :param _terminal_symbol_list: 终结符列表
    :param _itemSetFamily: 状态集
    :param _action: action表
    :param _goto: goto表
    :return: 推导过程输出到SLR_parsing_procedure变量，每行为(stack, symbols, input, action)；错误信息输出到SLR_parsing_error变量，每行为(line, column, error_token)
    """
    # 推导过程输出到SLR_parsing_procedure变量，每行为(stack, symbols, input, action)
    _SLR_parsing_procedure = list()
    # 错误信息输出到SLR_parsing_error变量，每行为(line, column, error_token)
    _SLR_parsing_error = list()
    # 初始化
    _stack = [0]  # 状态栈
    _symbols = ['$']  # 符号栈
    _symbols_value = [None]  # 符号栈的值
    _input = _tokens  # 待处理的输入符号
    _next_action = ''  # 下一步动作
    _input_index = 0  # _input到了_tokens的第几个
    # 开始推导
    while True:
        # 获取当前状态
        now_state = _stack[-1]
        # 获取当前输入
        now_input_token = _input[0]
        now_input_value = type_map.get(now_input_token.type, now_input_token.value)
        # 获取下一步动作
        if now_input_value in _terminal_symbol_list:
            _next_action = _action[now_state][now_input_value]
        elif now_input_value in _nonterminal_symbol_list:
            raise Exception(f"input中不可能直接出现非终结符(理论上不可能): {str(now_input_token)}")
        else:
            raise Exception(f"未知的Token(似乎没有在token预处理中过滤掉):{str(now_input_token)}")
        # 处理未定义的情况（错误恢复），并输出错误信息
        if _next_action == '':
            # 错误处理的思路如下：
            # 一、记录错误的行列号和错误的token，注意防止重复记录
            # 二、找到替换的错误序列：
            # 1. 假设symbols[i+1:]和input[:j+1]需要替换, i指针初始为len-1，j指针初始为0
            # 2. 指针j向后移直到input[j] in ["}", ";"]，注意：如果读取到了"{"那么必须匹配到一个"}"，两个"{"也要匹配两个"}"，以此类推
            # 2.1. 如果j已经到最后了还没有找到，那么判断是缺少;或}，或者function后多了一些不该有的符号。此时input的内容几乎没法分析，直接全部删除：j=len(input)-1，还要j-=1因为结尾有一个'$'
            # 3. 指针i向前移直到set(["block", "(closed_/open_)statement", "function", "program"] & goto[stack[i]] != set()。如果指针i前移过程中，在symbols[i]中遇到了这些非终结符中的某一个，那么匹配的非终结符优先级更低(防止死循环)
            # 3.1. 如果已经到了最前面还没匹配成功，那么判断是前文都有问题，直接全部删除，i=-1，i+=1，因为开头有一个'$'，匹配的goto操作为"program"
            # 3.2. 记录匹配的goto操作，匹配的优先级为："block" > "(closed_>open_)statement" > "function" > "program"
            # 三、输出错误信息：[ERROR] Reduce by (closed_/open_)statement/.../program -> symbols[i+1:] + input[:j+1]
            # 四、删除错误的序列，即stack=stack[:i], symbols=symbols[:i], input=input[j+1:]
            # 五、放入一个(closed_/open_)statement/.../program，就完成了错误序列的替换，然后执行对应的goto操作
            # 六、完成错误处理，执行continue
            pass
            # 一、记录错误的行列号和错误的token，注意防止重复记录
            if len(_SLR_parsing_error) == 0 or _SLR_parsing_error[-1] != \
                    [_token_place[_input_index][0], _token_place[_input_index][1], now_input_token.value]:
                _SLR_parsing_error.append(
                    [_token_place[_input_index][0], _token_place[_input_index][1], now_input_token.value])
            # 二、删除错误的序列
            # 1. 假设symbols[i+1:]和input[:j+1]需要替换, i指针初始为len-1，j指针初始为0
            i, j = len(_symbols) - 1, 0
            # 2. 指针j向后移直到input[j] in ["}", ";"]，注意：如果读取到了"{"那么必须匹配到一个"}"，两个"{"也要匹配两个"}"，以此类推
            left_brace_count = 0
            while j < len(_input):
                if _input[j].value == "{":
                    left_brace_count += 1
                elif _input[j].value == "}":
                    left_brace_count -= 1
                    if left_brace_count == 0:
                        break
                elif _input[j].value == ";" and left_brace_count == 0:
                    break
                j += 1
            # 2.1. 如果j已经到最后了还没有找到，那么判断是缺少;或}，或者function后多了一些不该有的符号。此时input的内容几乎没法分析，直接全部删除：j=len(input)-1，还要j-=1因为结尾有一个'$'
            if j >= len(_input) - 1:
                j = len(_input) - 1
                j -= 1  # 因为结尾有一个'$'
            # 3. 指针i向前移直到set(["block", "(closed_/open_)statement", "function", "program"] & goto[stack[i]] != set()。如果指针i前移过程中，在symbols[i]中遇到了这些非终结符中的某一个，那么匹配的非终结符优先级更低(防止死循环)
            # target_goto_key_list = ["Block", "Closed_statement", "Open_statement", "Statement", "Function", "Program"] # FIXME 删掉了部分
            target_goto_key_list = ["Block", "Stat", "Program"]
            goto_key = ""
            while i >= 0:
                goto_key_found = False
                # 3.2. 记录匹配的goto操作，匹配的优先级为："block" > "(closed_>open_)statement" > "function" > "program"
                for goto_key in target_goto_key_list:
                    if _goto[_stack[i]].get(goto_key, ""):
                        if not (j == -1 and i == len(_symbols) - 2 and target_goto_key_list.index(
                                _symbols[i + 1]) >= target_goto_key_list.index(goto_key)):
                            goto_key_found = True
                            break
                if goto_key_found:
                    break
                i -= 1
            # 3.1. 如果已经到了最前面还没匹配成功，那么判断是前文都有问题，直接全部删除，i=-1，i+=1，因为开头有一个'$'，匹配的goto操作为"Program"
            if i <= 0:
                i = 0
                goto_key = "Program"
            assert goto_key in target_goto_key_list, "未知的goto_key(理论上不可能)"
            # 三、输出错误信息：[ERROR] Reduce by (closed_/open_)statement/.../program -> symbols[i+1:] + input[:j+1]
            _next_action_str = f"[ERROR] Reduce by `{goto_key} -> {' '.join([str(i) for i in _symbols[i + 1:]])} {' '.join([type_map.get(i.type, i.value) for i in _input[:j + 1]])}`"
            _SLR_parsing_procedure.append([
                ' '.join([str(i) for i in _stack]),
                ' '.join([str(i) for i in _symbols]),
                ' '.join([str(type_map.get(i.type, i.value)) for i in _input]),
                _next_action_str
            ])
            # 四、删除错误的序列
            _stack, _symbols, _input = _stack[:i + 1], _symbols[:i + 1], _input[j + 1:]
            _input_index += j + 1
            # 五、放入一个(closed_/open_)statement/.../program，就完成了错误序列的替换，然后执行对应的goto操作
            _symbols.append(goto_key)
            now_state = _stack[-1]
            _stack.append(int(_goto[now_state][goto_key]))
            # 六、完成错误处理，执行continue
            continue

        # 判断action是否合法
        assert _next_action.count('s') + _next_action.count('r') < 2, "归约归约冲突或移入归约冲突(似乎之前没检查出来)"
        assert _next_action == 'acc' or (_next_action[0] in ['s', 'r'] and _next_action[1:].isnumeric()), \
            "未知的action表动作(似乎之前没检查出来)"
        # 动作转为对应字符串
        _next_action_str = ''
        if _next_action == 'acc':  # 在执行action时break
            _next_action_str = 'accept'
        elif _next_action[0] == 's' and _next_action[1:].isnumeric():
            _next_action_str = f'Shift to `{_next_action[1:]}`'
        elif _next_action[0] == 'r' and _next_action[1:].isnumeric():
            _next_action_str = f'''Reduce by `{
            _augmented_grammar_order_list[int(_next_action[1:])][0] + ' -> ' +
            ' '.join(_augmented_grammar_order_list[int(_next_action[1:])][1])
            }`'''
        else:
            raise Exception(f"未知的action表动作(似乎没有在上面处理掉): {_next_action}")
        # 输出到output_table
        _SLR_parsing_procedure.append([
            ' '.join([str(i) for i in _stack]),
            ' '.join([str(i) for i in _symbols]),
            ' '.join([str(type_map.get(i.type, i.value)) for i in _input]),
            _next_action_str
        ])

        # 执行action
        if _next_action == 'acc':
            break
        elif 's' in _next_action and _next_action[1:].isnumeric():
            _stack.append(int(_next_action[1:]))
            _symbols.append(now_input_value)
            _symbols_value.append(get_lex_value(now_input_token))
            _input = _input[1:]
            _input_index += 1
        elif 'r' in _next_action and _next_action[1:].isnumeric():
            reduce_index = int(_next_action[1:])
            reduce_production = _augmented_grammar_order_list[reduce_index]
            left_symbol, right_symbol_list = reduce_production
            popped_symbols_value = []
            for _ in range(len(right_symbol_list)):
                _stack.pop()
                _symbols.pop()  # 这个值也是符号值的栈它应该有的类型
                popped_symbols_value.append(_symbols_value.pop())
            popped_symbols_value.reverse()  # 出栈的顺序是反过来的，所以要反转
            left_symbol_value = get_reduce_value(left_symbol, right_symbol_list, popped_symbols_value,
                                                 nonterminal_symbol_list)
            _symbols.append(left_symbol)
            _symbols_value.append(left_symbol_value)
            now_state = _stack[-1]
            _stack.append(int(_goto[now_state][left_symbol]))
        else:
            raise Exception(f"未知的action表动作(似乎没有在上面处理掉): {_next_action}")

    return _SLR_parsing_procedure, _SLR_parsing_error


# %%
# 输出SLR推导过程到excel
def print_SLR_parsing_procedure_to_excel(_SLR_parsing_procedure: list[list[str]],
                                         column_names: tuple = ("Stack", "Symbols", "Input", "Action"),
                                         output_filepath: str = 'SLR_parsing_procedure.xlsx'):
    """
    输出SLR推导过程到excel
    :param _SLR_parsing_procedure: SLR推导过程
    :param column_names: 要输出的列名
    :param output_filepath: 输出文件路径
    """
    # 创建一个Excel工作簿
    wb = Workbook()
    # 选择默认的工作表
    ws = wb.active

    # 要输出的列名对应的顺序
    column_indexes = [["Stack", "Symbols", "Input", "Action"].index(column_name) + 1 for column_name in column_names]
    # 要输出的列名对应的宽度
    column_widths = [{"Stack": 10, "Symbols": 100, "Input": 80, "Action": 50}[column_name] for column_name in
                     column_names]

    # 设置表头
    ws.append(column_names)

    # 输出表体
    for row in _SLR_parsing_procedure:
        ws.append([row[i - 1] for i in column_indexes])

    # 设置每个单元格的为字符串格式
    for i in range(1, len(_SLR_parsing_procedure) + 2):
        for j in range(1, len(column_names) + 1):
            ws.cell(row=i, column=j).data_type = 's'

    # 设置每一列的对应宽度
    for i, width in enumerate(column_widths):
        column_letter = get_column_letter(i + 1)
        ws.column_dimensions[column_letter].width = width

    # 保存Excel文件
    wb.save(output_filepath)


# %%
# 预处理tokens
def preprocessing_tokens(_tokens: list[Token], _token_place: list[tuple[int, int]]) -> \
        tuple[list[Token], list[tuple[int, int]]]:
    """
    预处理tokens: 过滤掉不处理的类型：空白符、预处理命令、注释；末尾加一个结束符
    :param _tokens: 词法分析结果
    :param _token_place: token的位置(行号、列号)的列表
    :return: 预处理后的tokens
    """
    assert isinstance(_tokens, list) and len(_tokens) > 0, "词法分析结果为空，可能你的输入路径有问题"
    # 过滤掉不处理的类型：空白符、预处理命令、注释
    _token_place = [_token_place[i] for i in range(len(_token_place))
                    if _tokens[i].type not in ['whitespace', 'command', 'comment']]
    _tokens = list(filter(lambda _token: _token.type not in ['whitespace', 'command', 'comment'], _tokens))
    # 报错，不该有的类型：错误
    assert len(list(filter(lambda _token: _token.type in ['error'], _tokens))) == 0, \
        f"不该有的类型: error，你的词法分析未通过，请检查代码的词法错误，可以使用上次的可视化"
    # 报错，暂不支持的运算符
    for token in _tokens:
        if token.type == 'operator' and token.value not in \
                ['+', '+=', '-', '-=', '*', '*=', '/', '/=', '%', '%=', '!', '!=',
                 '<', '<=', '=', '==', '>', '>=',
                 '&&', '||', '&',
                 '(', ')', '[', ']', '{', '}', ';', ',']:
            raise Exception(f"暂不支持的运算符: {str(token)}")
    # 报错，暂不支持的关键字
    for token in _tokens:
        if token.type == 'keyword' and token.value not in \
                ['int', 'float', 'char', 'string', 'bool', 'void',
                 'long', 'double',
                 'true', 'false',
                 'if', 'else', 'while', 'do', 'for', 'goto', 'break', 'continue',
                 'return']:
            raise Exception(f"暂不支持的关键字: {str(token)}")
    # 末尾加一个结束符
    _token_place.append((-1, -1))
    _tokens.append(Token('end', '$', ''))

    return _tokens, _token_place


# %%
# 获取token的位置(行号、列号)
def get_token_place(_tokens: list[Token]) -> list[tuple[int, int]]:
    """
    获取token的位置(行号、列号)的列表，用于报错时提示错误位置，通过token_place[token_index]来使用
    :param _tokens: 词法分析结果，注意：此处需要过滤前的tokens
    :return: token的位置的列表(行号、列号)，查找时用token_place[token_index]
    """
    _token_place = list()
    _line_number = 1
    _column_number = 1
    for token in _tokens:
        _token_place.append((_line_number, _column_number))
        for char in token.value:
            if char == '\n':
                _line_number += 1
                _column_number = 1
            else:
                _column_number += 1
    return _token_place


def get_lex_value(_token: Token) -> my_sdt_action.Symbol.id | my_sdt_action.Symbol.num | None:
    """
    获取token的文法值，处理int和float值，id，其他返回None，类型似乎也要返回None
    :param _token: token
    :return: token的值
    """
    if _token.type in ['int_dec', 'int_oct', 'int_bin', 'int_hex']:  # int数值
        return my_sdt_action.Symbol.num(int(_token.value, base=10), lextype=my_sdt_action.int)
    elif _token.type == 'float':  # float数值
        return my_sdt_action.Symbol.num(float(_token.value), lextype=my_sdt_action.float)
    elif _token.type == 'identifier':  # id
        return my_sdt_action.Symbol.id(_token.value)
    # elif _token.type == 'keyword' and _token.value in ['int', 'long', 'float', 'double']:  # 类型
    #     return my_sdt_action.Symbol
    else:
        return None


def get_reduce_value(_left_symbol: str,
                     _right_symbol_list: my_sdt.SDT_right[str], _popped_symbols_value: list[my_sdt_action.Symbol],
                     _nonterminal_symbol_list: list[str]) -> \
        my_sdt_action.Symbol._symbol | None:
    """
    获取归约的值（其实也是执行对应语义动作）
    :param _left_symbol: 归约的左侧符号
    :param _right_symbol_list: 归约的右侧符号列表
    :param _popped_symbols_value: 归约的右侧符号值列表
    :param _nonterminal_symbol_list: 非终结符列表，用于判断是否有值
    :return: 归约的值
    """
    # 步骤如下
    # 产生每个右部符号对应的值的变量(有值的才生成变量)
    # 看看需不需要引入my_sdt的内部变量
    # 再产生一个归约的值的实例，用来存储结果
    # 执行action的代码
    # 返回值

    try:
        # 定义一个local_context，用于存储局部变量
        local_context = dict()

        # 产生每个右部符号对应的值的变量(有值的才生成变量)
        for i in range(len(_right_symbol_list)):
            # 断言两种判断方式相同
            assert (_popped_symbols_value[i] is not None) \
                   == (_right_symbol_list[i] in ['num',
                                                 'id'] + _nonterminal_symbol_list), "两种判断方式不同, 可能是符号栈中的非终结符的值没有正确赋值"
            if _popped_symbols_value[i] is not None:
                local_context[_right_symbol_list.rights[i]] = _popped_symbols_value[i]  # 注意这里要用.rights[i]，而不是[i]

        # 看看需不需要引入my_sdt的内部变量
        inner_variable = [
            'nextinstr', 'top', '_error_list',
            'backpatch', 'merge', 'makelist', 'gen', 'newTemp', 'array', 'error', 'resultAccuracy',
            'int', 'long', 'float', 'double', ]
        for i in inner_variable:
            if i in _right_symbol_list.a:
                exec(f"from my_sdt_action import {i}", globals(), local_context)
        # 再产生一个归约的值的实例，用来存储结果
        exec(f"{_left_symbol} = my_sdt_action.Symbol.{_left_symbol}()", globals(), local_context)
        # 执行action的代码
        # exec(_right_symbol_list.a) # 应该分行执行（且去""""""长字符串所带来的空格，这个在导入的包中处理）
        exec(_right_symbol_list.a, globals(), local_context)
        # 返回值
        result = local_context.get(_left_symbol)
        return result
    except Exception as e:
        print(f"执行{_left_symbol} -> {' '.join(_right_symbol_list.rights)}的action时出错: {e}")
        raise e


# 输出sdt_grammar
def print_sdt_grammar(sdt_grammar):
    for left_symbol in sdt_grammar.keys():
        for production_index in range(len(sdt_grammar[left_symbol])):
            if len(sdt_grammar[left_symbol][production_index]) == 0:
                print(f"{left_symbol} -> ε", end="")
            else:
                print(f"{left_symbol} -> {' '.join(sdt_grammar[left_symbol][production_index].rights)}", end="")
            if sdt_grammar[left_symbol][production_index].a == 'pass\n':
                print()
            else:
                indent_space = 4 * " "
                # a[:-1]去掉最后一个换行符
                print(("\n" + indent_space).join(
                    [""] + ("{" + sdt_grammar[left_symbol][production_index].a[:-1] + "}").split("\n")
                ))
            print()


# 输出三地址代码
def print_three_address_code(_result_code: list[str]):
    """
    输出三地址代码
    :param _result_code: 三地址代码
    """
    for line_num, line in enumerate(_result_code):
        print(f"{line_num}: {line}")


# 输出语义错误列表
def print_error_list(_error_list: list[str]):
    """
    输出语义错误列表
    :param _error_list: 语义错误列表
    """
    for error in _error_list:
        print(f"{error}")


# %%
# 主函数
if __name__ == '__main__':
    for file_id in ["1"]:
        # 定义文件路径
        input_filepath = rf'./data/input/{file_id}_sample.c'  # 输入文件
        output_SLR_parsing_table_filepath = rf'./data/output/{file_id}_SLR_parsing_table.txt'  # 输出SLR分析表文件
        output_SLR_parsing_table_excelPath = rf'./data/output/{file_id}_SLR_parsing_table.xlsx'  # 输出SLR分析表表格
        output_SLR_parsing_procedure_all_excelPath = rf'./data/output/{file_id}_SLR_parsing_procedure(all).xlsx'  # 输出SLR分析过程表格(全部)
        output_SLR_parsing_procedure_symbolAndAction_excelPath = rf'./data/output/{file_id}_SLR_parsing_procedure(symbol action).xlsx'  # 输出SLR分析过程表格(符号和动作)
        output_SLR_parsing_error_filepath = rf'./data/output/{file_id}_SLR_parsing_error.txt'  # 输出SLR分析错误文件
        output_SDT_file = rf'./data/output/{file_id}_SDT_file.txt'  # 输出SDT文件
        output_SDT_result_3code_filepath = rf'./data/output/{file_id}_SDT_result_3code.txt'  # 输出SDT结果三地址代码文件
        output_SDT_result_error_filepath = rf'./data/output/{file_id}_SDT_result_error.txt'  # 输出SDT结果错误文件
        # 检查输入文件是否存在
        assert os.path.exists(input_filepath), f"输入文件不存在: {input_filepath}"
        # 如果输出目录./data/output则创建文件夹
        if not os.path.exists('./data/output'):
            os.makedirs('./data/output')
        # 刷新输出文件('w'模式会清空文件)
        print_redirect_builder(output_SLR_parsing_table_filepath, use_cache=False)("", end="", mode='w')
        print_redirect_builder(output_SLR_parsing_error_filepath, use_cache=False)("", end="", mode='w')

        # 导入词法分析器
        from task1_package import Token, get_token

        # 获取词法分析结果
        tokens = get_token(input_filepath)
        # 定义类型映射，避免混淆。将下面的这些类型的Token映射为对应的非终结符，而其他类型直接用value当作非终结符
        # （本质为调用type_map.get(token.type, token.value)，如token(1)会映射为'int_value', token(;)则不变';'）
        type_map = {
            'int_dec': 'num',
            'int_oct': 'num',
            'int_bin': 'num',
            'int_hex': 'num',
            'float': 'num',
            # 'char': 'char_value',
            # 'string': 'string_value',
            'identifier': 'id',
        }
        # 获取token的位置(行号、列号)，用于语法分析报错
        token_place = get_token_place(tokens)
        # 预处理tokens：过滤掉不处理的类型：空白符、预处理命令、注释；末尾加一个结束符
        tokens, token_place = preprocessing_tokens(tokens, token_place)

        # 读取文法
        # from my_sdt import sdt_grammar as augmented_grammar
        from my_sdt import sdt_grammar as augmented_grammar

        # 检查文法格式
        check_grammar(augmented_grammar)
        # 获取符号列表
        symbol_list, nonterminal_symbol_list, terminal_symbol_list = get_symbol_list(augmented_grammar)
        # 定义文法顺序
        augmented_grammar_to_index, augmented_grammar_order_list = get_augmented_grammar_to_index(augmented_grammar)
        # 获取First集和Follow集
        first = get_first(augmented_grammar, nonterminal_symbol_list, terminal_symbol_list)
        follow = get_follow(augmented_grammar, nonterminal_symbol_list, terminal_symbol_list, first)
        # 获取状态集
        itemSetFamily = get_itemSetFamily(augmented_grammar, symbol_list)
        # 获取SLR分析表
        action, goto = get_SLR_table(
            augmented_grammar, augmented_grammar_to_index,
            symbol_list, nonterminal_symbol_list, terminal_symbol_list,
            first, follow, itemSetFamily)
        # 输出SLR分析表到文本文件
        print_SLR_table(nonterminal_symbol_list, terminal_symbol_list, itemSetFamily, action, goto,
                        output_SLR_parsing_table_filepath)
        # 输出SLR分析表到Excel文件
        print_SLR_table_to_excel(nonterminal_symbol_list, terminal_symbol_list, itemSetFamily, action, goto,
                                 output_SLR_parsing_table_excelPath)
        # 检查SLR冲突
        check_SLR(action, goto)
        # 开始推导
        SLR_parsing_procedure, SLR_parsing_error = SLR_parsing(
            tokens, token_place,
            augmented_grammar_order_list, nonterminal_symbol_list, terminal_symbol_list,
            itemSetFamily, action, goto)
        # 输出错误信息
        print = print_redirect_builder(output_SLR_parsing_error_filepath)
        for error in SLR_parsing_error:
            print(f"[ERROR] Line {error[0]}, Column {error[1]}: Unexpected token `{error[2]}`")
        print("\n", flush=True)
        # 输出推导过程(all)到excel
        print_SLR_parsing_procedure_to_excel(SLR_parsing_procedure, ("Stack", "Symbols", "Input", "Action"),
                                             output_SLR_parsing_procedure_all_excelPath)
        # 输出推导过程(仅符号和动作)到excel
        print_SLR_parsing_procedure_to_excel(SLR_parsing_procedure, ("Symbols", "Action"),
                                             output_SLR_parsing_procedure_symbolAndAction_excelPath)

        # 获取语义分析结果
        from my_sdt_action import _result_code, _error_list

        # 输出sdt语法
        print = print_redirect_builder(output_SDT_file, use_cache=False)
        print_sdt_grammar(augmented_grammar)
        # 输出三地址代码
        print = print_redirect_builder(output_SDT_result_3code_filepath, use_cache=False)
        print_three_address_code(_result_code)
        # 输出语义错误列表
        print = print_redirect_builder(output_SDT_result_error_filepath, use_cache=False)
        print_error_list(_error_list)

        # 恢复输出，防止后续输出被重定向
        print = builtins.print
